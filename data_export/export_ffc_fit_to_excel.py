#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Sep  2 22:03:17 2024

@author: Joachim
"""

# import dependencies
import numpy as np
import os

#import local dependencies
from math_functions.mapping_functions import all_Dispersion_functions


# Class to do the exporting of the FFC analysis to excel
def ExportFFCFitToExcel(ExportPath, ExportDataSet):
    # import pandas as pd
    import openpyxl as opyxl
    
    def GetFitVariableNames(FitType):
        import inspect
        tmpVariables = str(inspect.signature(all_Dispersion_functions(FitType)))
        tmpVariables = tmpVariables.strip('(')
        tmpVariables = tmpVariables.strip(')')
        tmpVariables = tmpVariables.split(',')
        
        Vars_out = []
        for var_out in tmpVariables:
            Vars_out.append(var_out.strip(' '))
        return Vars_out
        
    # Define font types
    Header_font = opyxl.styles.Font(name='Calibri',size=14,bold=True)
    Excluded_Dispersion_Value_font = opyxl.styles.Font(color="FF0000")
    
    # Create an excel workbook 
    wb = opyxl.Workbook()
    
    # Create a Summary sheet in the workbook to export all the fit parameters into
    ws_Summary = wb.create_sheet('Summary',0)
    
    
    ws_Summary['A1'] = "Dispersion Fit Type:"
    ws_Summary['A1'].font = Header_font
    
    ws_Summary['B1'] = ExportDataSet[0].disp_fit[0]
    ws_Summary['B1'].font = Header_font
    
    
    FitVars = GetFitVariableNames(ExportDataSet[0].disp_fit[0])
    FitVars.remove(FitVars[0])
    
    for count, FVar in enumerate(FitVars):
        ws_Summary.cell(row=1,column=count+3).value = FVar
        ws_Summary.cell(row=1,column=count+3).font = Header_font
    
    # Add R-squared values
    ws_Summary.cell(row=1,column=len(FitVars)+4).value = "R-squared:"
    ws_Summary.cell(row=1,column=len(FitVars)+4).font = Header_font
    
    
    ws_Summary.cell(row=1,column=len(FitVars)+6).value = "Parameter Variance:"
    ws_Summary.cell(row=1,column=len(FitVars)+6).font = Header_font
    
    for count, FVar in enumerate(FitVars):
        ws_Summary.cell(row=1,column=count+len(FitVars)+7).value = FVar
        ws_Summary.cell(row=1,column=count+len(FitVars)+7).font = Header_font    
    
    # Define offset parameter in order to adapt for multiple dispersion profiles per sample
    MultiDispRowOffset = 0
    
    # Loop to export all the fit parameters and their variances
    for count, Sample in enumerate(ExportDataSet):
        
        # Export the sample name and format in header font
        strSampleName = str(Sample.file_name)
        ws_Summary.cell(row = count + MultiDispRowOffset + 2,column=2).value = strSampleName
        ws_Summary.cell(row = count + MultiDispRowOffset + 2,column=2).font = Header_font
        
        
        # Export R-squared of the present fit
        ws_Summary.cell(row = count + MultiDispRowOffset + 2,column=len(FitVars)+4).value = Sample.disp_fit[4]
        
        # Export all Parameters used in present fit
        for count_var, FVar in enumerate(Sample.disp_fit[1]):
            ws_Summary.cell(row = count + MultiDispRowOffset + 2,column=count_var+3).value = FVar
        
        # Export the variance of the fitted parameters
        for count_var, FVar in enumerate(Sample.disp_fit[3]):
            # Check for inf values as these would result n empty entries in excel instead write inf as a string into these cells
            if np.isfinite(FVar):
                ws_Summary.cell(row = count + MultiDispRowOffset + 2,column=count_var+len(FitVars)+7).value = FVar
            else:
                ws_Summary.cell(row = count + MultiDispRowOffset + 2,column=count_var+len(FitVars)+7).value = "inf"
                
                
        # Check if more than one Dispersionfit exists for the same sample
        if len(Sample.disp_fit) == 10:
            # Increase counter to keep track of multiple rows
            MultiDispRowOffset += 1
            
            #Export the sample name and add _2nd to the end of the name
            strSampleName = strSampleName + "_2nd"
            ws_Summary.cell(row=count + MultiDispRowOffset + 2,column=2).value = strSampleName
            ws_Summary.cell(row=count + MultiDispRowOffset + 2,column=2).font = Header_font
            
            # Export R-squared of the present fit
            ws_Summary.cell(row = count + MultiDispRowOffset + 2,column=len(FitVars)+4).value = Sample.disp_fit[9]
            
            # Export all Parameters used in present fit
            for count_var, FVar in enumerate(Sample.disp_fit[6]):
                ws_Summary.cell(row = count + MultiDispRowOffset + 2,column=count_var+3).value = FVar
            
            # Export the variance of the fitted parameters
            for count_var, FVar in enumerate(Sample.disp_fit[8]):
                # Check for inf values as these would result n empty entries in excel instead write inf as a string into these cells
                if np.isfinite(FVar):
                    ws_Summary.cell(row = count + MultiDispRowOffset + 2,column=count_var+len(FitVars)+7).value = FVar
                else:
                    ws_Summary.cell(row = count + MultiDispRowOffset + 2,column=count_var+len(FitVars)+7).value = "inf"
                    
    # Merge rows where sample data was written to in order to create space for a title
    ws_Summary.merge_cells(start_row=2,start_column=1,end_column=1,end_row=len(ExportDataSet) + MultiDispRowOffset + 1)    
    ws_Summary['A2'] = "Sample Names"
    ws_Summary['A2'].font = Header_font
    ws_Summary['A2'].alignment = opyxl.styles.Alignment(text_rotation=90,wrap_text=True)    
    
    ## Final formatting cleanups
    # Iterate over all columns and adjust their widths
    for column in ws_Summary.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws_Summary.column_dimensions[column_letter].width = adjusted_width
     
    ## Export the dispersion fields on an extra sheet of the excel file           
    # Create Dispersion sheet
    ws_Dispersions = wb.create_sheet('Dispersion',1)
    
    # Add header to the sheet
    ws_Dispersions['A1'] = "Relaxation Field BR:"
    ws_Dispersions['A1'].font = Header_font
    
    # Use the first dataset to determine the field strengths and export them    
    for count, FVar in enumerate(ExportDataSet[0].zone_parameters['BR']):
        ws_Dispersions.cell(row=1,column=count+3).value = str(float(FVar)*1e6)
        ws_Dispersions.cell(row=1,column=count+3).font = Header_font
    
    # Define offset parameter in order to adapt for multiple dispersion profiles per sample
    MultiDispRowOffset = 0
    
    # Export the dispersion data
    for count, Sample in enumerate(ExportDataSet):
        
        # Export the sample name and format in header font
        strSampleName = str(Sample.file_name)
        ws_Dispersions.cell(row = count + MultiDispRowOffset + 2, column = 2).value = strSampleName
        ws_Dispersions.cell(row = count + MultiDispRowOffset + 2, column = 2).font = Header_font
        
        # Lookup the mask of the current dataset
        Mask = Sample.zone_parameters['DataMask'][0]
        
        # Loop to export all the dispersion field values
        for count_var, FVar in enumerate(Sample.zone2disp_data[0]):
            ws_Dispersions.cell(row = count + MultiDispRowOffset + 2, column = count_var + 3).value = FVar
            
            # Check wether current dispersion value has been masked for fit and change font if so
            if Mask[count_var] == False:
                ws_Dispersions.cell(row = count + MultiDispRowOffset + 2, column = count_var + 3).font = Excluded_Dispersion_Value_font
            
        # Check if more than one Dispersion profile exists for the same sample
        if len(Sample.zone2disp_data) == 6:
            # Increase counter to keep track of multiple rows
            MultiDispRowOffset += 1
            
            #Export the sample name and add _2nd to the end of the name
            strSampleName = strSampleName + "_2nd"
            ws_Dispersions.cell(row = count + MultiDispRowOffset + 2, column = 2).value = strSampleName
            ws_Dispersions.cell(row = count + MultiDispRowOffset + 2, column = 2).font = Header_font
            
            # Lookup the mask of the current dataset
            Mask = Sample.zone_parameters['DataMask'][1]
            
            # Loop to export all the dispersion field values
            for count_var, FVar in enumerate(Sample.zone2disp_data[4]):
                ws_Dispersions.cell(row = count + MultiDispRowOffset + 2, column = count_var + 3).value = FVar
                
                # Check wether current dispersion value has been masked for fit and change font if so
                if Mask[count_var] == False:
                    ws_Dispersions.cell(row = count + MultiDispRowOffset + 2, column = count_var + 3).font = Excluded_Dispersion_Value_font
                
                
    # Merge rows where sample data was written to in order to create space for a title
    ws_Dispersions.merge_cells(start_row=2,start_column=1,end_column=1,end_row=len(ExportDataSet) + MultiDispRowOffset + 1)
    ws_Dispersions['A2'] = "Sample Names"
    ws_Dispersions['A2'].font = Header_font
    ws_Dispersions['A2'].alignment = opyxl.styles.Alignment(text_rotation=90,wrap_text=True)
    
    ## Final formatting cleanups
    # Iterate over all columns and adjust their widths
    for column in ws_Dispersions.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws_Dispersions.column_dimensions[column_letter].width = adjusted_width
    
    # Save the created excel file in the folder where the first data set was read from
    wb.save(os.path.join(ExportPath, 'FFC Analysis.xlsx') )  